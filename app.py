from flask import Flask, render_template, request, redirect
import sqlite3
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

    # ==========================================
    # CREAR BASE DE DATOS
    # ==========================================
def crear_db():
    conn = sqlite3.connect("visitas.db")
    cursor = conn.cursor()
 
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS visitas (
 
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT,
        nombre TEXT,
        empresa TEXT,
        ine TEXT,
        persona_visita TEXT,
        motivo TEXT,
        hora_entrada TEXT,
        hora_salida TEXT,
        foto_visitante TEXT
 
        )
    """)
 
    conn.commit()
    conn.close()

crear_db()

def guardar_excel():
    
        fecha_excel = datetime.now().strftime("%Y-%m-%d")
 
 
    # ==========================================
    # GUARDAR EN EXCEL
    # ==========================================
def guardar_excel():
        fecha_excel = datetime.now().strftime("%Y-%m-%d")
 
        if not os.path.exists("Historial"):
            os.makedirs("Historial")
 
        archivo = f"Historial/Historial_{fecha_excel}.xlsx"
 
        if not os.path.exists(archivo):
 
            wb = Workbook()
            ws = wb.active
 
            ws.title = "Visitas"
 
            ws.append([
                "ID",
                "Fecha",
                "Nombre",
                "Empresa",
                "INE",
                "Persona",
                "Motivo",
                "Entrada",
                "Salida"
            ])
 
            wb.save(archivo)
 
        wb = load_workbook(archivo)
        ws = wb.active
 
        conn = sqlite3.connect("visitas.db")
        cursor = conn.cursor()
 
        fecha_hoy = datetime.now().strftime("%d/%m/%Y")
 
        cursor.execute("""
            SELECT *
            FROM visitas
            WHERE fecha = ?
        """, (fecha_hoy,))
 
        visitas = cursor.fetchall()
 
        conn.close()
 
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
 
        for visita in visitas:
 
            ws.append([
                visita[0],
                visita[1],
                visita[2],
                visita[3],
                visita[4],
                visita[5],
                visita[6],
                visita[7],
                visita[8]
            ])
 
        wb.save(archivo)


# ==========================================
# PAGINA PRINCIPAL
# ==========================================

@app.route("/")
def index():

    conn = sqlite3.connect("visitas.db")
    cursor = conn.cursor()

    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
 
    cursor.execute("""
        SELECT *
        FROM visitas
        WHERE fecha = ?
        ORDER BY id DESC
    """, (fecha_hoy,))
 
    visitas = cursor.fetchall()
 
    conn.close()
 
    return render_template("index.html", visitas=visitas)

# ==========================================
# REGISTRAR VISITA
# ==========================================

@app.route("/registrar", methods=["POST"])
def registrar():

    fecha = datetime.now().strftime("%d/%m/%Y")
    hora_entrada = datetime.now().strftime("%H:%M:%S")

    nombre = request.form.get("nombre")
    empresa = request.form.get("empresa")

    ine_completo = request.form.get("ine")

    ine = ""

    if ine_completo:

        numeros = ''.join(filter(str.isdigit, ine_completo))
        ine = numeros[:13]

    persona = request.form.get("persona")
    motivo = request.form.get("motivo")
    foto_visitante = request.form.get("foto_visitante")

    conn = sqlite3.connect("visitas.db")
    cursor = conn.cursor()

    cursor.execute("""

    INSERT INTO visitas (

        fecha,
        nombre,
        empresa,
        ine,
        persona_visita,
        motivo,
        hora_entrada,
        hora_salida,
        foto_visitante

    )

    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)

    """, (

        fecha,
        nombre,
        empresa,
        ine,
        persona,
        motivo,
        hora_entrada,
        "",
        foto_visitante

    ))

    conn.commit()
    guardar_excel()
    conn.close()

    return redirect("/")

# ==========================================
# ELIMINAR REGISTRO
# ==========================================

@app.route("/eliminar/<int:id>")
def eliminar(id):

    conn = sqlite3.connect("visitas.db")
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM visitas WHERE id = ?",
        (id,)
    )

    conn.commit()
    guardar_excel()
    conn.close()

    return redirect("/")

# ==========================================
# REGISTRAR SALIDA
# ==========================================

@app.route("/salida/<int:id>")
def salida(id):

    hora_salida = datetime.now().strftime("%H:%M:%S")

    conn = sqlite3.connect("visitas.db")
    cursor = conn.cursor()

    cursor.execute("""

        UPDATE visitas
        SET hora_salida = ?
        WHERE id = ?

    """, (hora_salida, id))

    conn.commit()
    guardar_excel()
    conn.close()

    return redirect("/")

# ==========================================
# IMPRIMIR ETIQUETA
# ==========================================

@app.route("/imprimir/<int:id>")
def imprimir(id):

    conn = sqlite3.connect("visitas.db")
    cursor = conn.cursor()

    cursor.execute("""

        SELECT *
        FROM visitas
        WHERE id = ?

    """, (id,))

    visita = cursor.fetchone()

    conn.close()

    foto_html = ""

    if visita[9]:

        foto_html = f"""
            <img src="{visita[9]}">
        """

    return f"""

    <html>

    <head>

        <title>Etiqueta</title>

        <style>

            @page {{
                size: 10cm 5.8cm landscape;
                margin: 0;
            }}

            body {{

                margin: 0;
                padding: 0;

                width: 10cm;
                height: 5.8cm;

                overflow: hidden;

                font-family: Arial, sans-serif;

            }}

            .etiqueta {{

                width: 10cm;
                height: 5.8cm;

                box-sizing: border-box;

                display: flex;
                flex-direction: row;

                justify-content: space-between;
                align-items: flex-start;

                padding-top: 4mm;
                padding-left: 7mm;
                padding-right: 4mm;

            }}

            .izquierda {{

                width: 58%;

            }}

            .titulo {{

                font-size: 18px;
                font-weight: bold;

                line-height: 1;

                margin-bottom: 4px;

            }}

            .nombre {{

                font-size: 24px;
                font-weight: bold;

                line-height: 1;

                margin-bottom: 8px;

                white-space: nowrap;

            }}

            .info {{

                font-size: 14px;
                margin-bottom: 3px;

                line-height: 1.1;

            }}

            .derecha {{

                width: 38%;

                display: flex;
                justify-content: center;
                align-items: flex-end;

                padding-top: 18mm;
                padding-right: 2mm;

            }}

            .derecha img {{

                width: 145px;
                height: 110px;

                object-fit: cover;

                border: 2px solid black;

            }}

        </style>

    </head>

    <body onload="window.print(); setTimeout(() => window.close(), 500)">

        <div class="etiqueta">

            <div class="izquierda">

                <div class="titulo">
                    NORDSON MEDICAL
                </div>

                <div class="nombre">
                    {visita[2]}
                </div>

                <div class="info">
                    Empresa: {visita[3]}
                </div>

                <div class="info">
                    INE: {visita[4]}
                </div>

                <div class="info">
                    Visita a: {visita[5]}
                </div>

                <div class="info">
                    Entrada: {visita[7]}
                </div>

                <div class="info">
                    Fecha: {visita[1]}
                </div>

            </div>

            <div class="derecha">

                {foto_html}

            </div>

        </div>

    </body>

    </html>

    """

# ==========================================
# INICIAR SERVIDOR
# ==========================================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000,debug=True)